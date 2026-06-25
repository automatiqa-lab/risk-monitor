"""
EML Loader - parses .eml email files and extracts structured intelligence
from the email body, PDF attachments, and XLSX attachments.

Converts everything into Article objects that feed into the agent pipeline.
Handles:
  - Email body text (equipment stock tables, free-form intelligence)
  - PDF attachments (fuel price reports, vessel schedules)
  - XLSX attachments (carrier advisory summaries with dates/details/regions)
"""
from __future__ import annotations

import email as email_lib
import logging
import re
import tempfile
from datetime import datetime, timezone
from email import policy
from pathlib import Path
from typing import List, Optional

from agent.rss_aggregator import Article

logger = logging.getLogger(__name__)


# ── Public API ───────────────────────────────────────────────────────────────

def load_eml(eml_path: Path) -> List[Article]:
    """
    Parse an .eml file and return Article objects extracted from:
      1. The email body (plain text)
      2. PDF attachments (text extraction via PyMuPDF)
      3. XLSX attachments (structured row-per-advisory parsing)

    Args:
        eml_path: Path to the .eml file.

    Returns:
        List of Article objects ready for the pipeline.
    """
    with open(eml_path, "rb") as f:
        msg = email_lib.message_from_binary_file(f, policy=policy.default)

    subject = msg.get("subject", "")
    sender = msg.get("from", "")
    date_str = msg.get("date", "")
    msg_date = _parse_email_date(date_str)

    logger.info("Parsing EML: %s (from: %s)", subject, sender)

    articles: List[Article] = []

    # 1. Email body
    body_text = _extract_body_text(msg)
    if body_text:
        body_articles = _parse_body_articles(body_text, subject, sender, msg_date)
        articles.extend(body_articles)
        logger.info("  Body: %d article(s) extracted", len(body_articles))

    # 2. Attachments
    for part in msg.walk():
        filename = part.get_filename()
        if not filename:
            continue

        payload = part.get_payload(decode=True)
        if not payload:
            continue

        lower_fn = filename.lower()

        if lower_fn.endswith(".pdf"):
            pdf_articles = _parse_pdf_attachment(payload, filename, msg_date)
            articles.extend(pdf_articles)
            logger.info("  PDF '%s': %d article(s)", filename, len(pdf_articles))

        elif lower_fn.endswith(".xlsx"):
            xlsx_articles = _parse_xlsx_attachment(payload, filename, msg_date)
            articles.extend(xlsx_articles)
            logger.info("  XLSX '%s': %d article(s)", filename, len(xlsx_articles))

    logger.info("EML total: %d articles from '%s'", len(articles), eml_path.name)
    return articles


def load_all_eml(input_dir: Path) -> List[Article]:
    """Load all .eml files from a directory."""
    articles: List[Article] = []
    for eml_path in sorted(input_dir.glob("*.eml")):
        try:
            articles.extend(load_eml(eml_path))
        except Exception as exc:
            logger.error("Failed to parse EML '%s': %s", eml_path.name, exc)
    return articles


# ── Email body parsing ───────────────────────────────────────────────────────

def _extract_body_text(msg) -> str:
    """Extract plain text body from email message."""
    for part in msg.walk():
        if part.get_content_type() == "text/plain":
            payload = part.get_payload(decode=True)
            if payload:
                return payload.decode("utf-8", errors="replace")
    return ""


def _parse_body_articles(
    body: str, subject: str, sender: str, msg_date: datetime
) -> List[Article]:
    """
    Parse the email body for structured intelligence.
    Extracts:
      - Equipment stock tables (carrier: 20ft - N, 40ft - N)
      - Free-form paragraphs with logistics intelligence
    """
    articles: List[Article] = []

    # Clean body: strip signatures, disclaimers
    body = _strip_email_noise(body)
    if not body.strip():
        return articles

    # Try to extract equipment stock table
    stock_article = _extract_equipment_stock(body, subject, sender, msg_date)
    if stock_article:
        articles.append(stock_article)

    # Extract intelligence paragraphs (non-stock content)
    paragraphs = _extract_intelligence_paragraphs(body, subject, sender, msg_date)
    articles.extend(paragraphs)

    return articles


def _extract_equipment_stock(
    body: str, subject: str, sender: str, msg_date: datetime
) -> Optional[Article]:
    """
    Look for carrier equipment stock patterns like:
        OOCL:
        20' - 123
        40' - 78
    Consolidate into a single summary article.
    """
    # Pattern: carrier name followed by 20'/40' stock lines
    carrier_pattern = re.compile(
        r"^([A-Z][A-Za-z\s]{1,25}):?\s*$",
        re.MULTILINE,
    )
    stock_pattern = re.compile(
        r"(20|40)['\u2018\u2019\u201C\u201D]?\s*[-\u2013\u2014]\s*(.+)",
        re.IGNORECASE,
    )

    carriers_found = []
    lines = body.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        carrier_match = carrier_pattern.match(line)
        if carrier_match:
            carrier_name = carrier_match.group(1).strip()
            # Check if next lines contain stock data
            stock_lines = []
            j = i + 1
            while j < len(lines) and j <= i + 5:
                sline = lines[j].strip()
                if not sline:
                    j += 1
                    continue
                sm = stock_pattern.match(sline)
                if sm:
                    stock_lines.append(sline)
                    j += 1
                else:
                    break
            if stock_lines:
                carriers_found.append(f"{carrier_name}: {'; '.join(stock_lines)}")
                i = j
                continue
        i += 1

    if not carriers_found:
        return None

    summary = "Equipment stock levels - " + " | ".join(carriers_found)
    return Article(
        title=f"Equipment Update: {subject}",
        url="",
        source=_extract_sender_name(sender),
        published_date=msg_date,
        raw_text=summary,
        summary=summary,
        regions=_infer_regions_from_subject(subject),
        container_signal="shortage" if any(
            kw in body.lower() for kw in ("low", "depleted", "nil", "shortage", "0\n", "zero")
        ) else "general",
    )


def _extract_intelligence_paragraphs(
    body: str, subject: str, sender: str, msg_date: datetime
) -> List[Article]:
    """Extract meaningful intelligence paragraphs from the email body."""
    articles: List[Article] = []

    # Split into paragraphs, filter short/noise
    paragraphs = re.split(r"\n\s*\n", body)
    intel_keywords = (
        "surcharge", "capacity", "transport", "congestion", "delay",
        "shortage", "strike", "conflict", "fuel", "equipment", "vessel",
        "booking", "holiday", "suspended", "monitor", "challenging",
    )

    for para in paragraphs:
        para = para.strip()
        if len(para) < 40 or len(para) > 1500:
            continue
        # Skip if it looks like a signature or disclaimer
        if any(kw in para.lower() for kw in ("disclaimer", "confidential", "regards", "@", "www.")):
            continue
        # Keep if it contains logistics intelligence keywords
        lower = para.lower()
        if any(kw in lower for kw in intel_keywords):
            articles.append(Article(
                title=f"Intel: {para[:80]}...",
                url="",
                source=_extract_sender_name(sender),
                published_date=msg_date,
                raw_text=para,
                summary=para[:300],
                regions=_infer_regions_from_subject(subject),
                container_signal=_infer_signal(para),
            ))

    return articles


# ── PDF attachment parsing ───────────────────────────────────────────────────

def _parse_pdf_attachment(
    payload: bytes, filename: str, msg_date: datetime
) -> List[Article]:
    """Extract text from PDF and convert to articles."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.warning("PyMuPDF not installed - skipping PDF '%s'", filename)
        return []

    articles: List[Article] = []

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(payload)
        tmp_path = tmp.name

    try:
        doc = fitz.open(tmp_path)
        full_text = ""
        for page in doc:
            full_text += page.get_text() + "\n"
        doc.close()
    except Exception as exc:
        logger.error("Failed to read PDF '%s': %s", filename, exc)
        return []
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    if not full_text.strip():
        return []

    # Determine the type of PDF content
    lower = full_text.lower()
    clean_name = filename.replace(".pdf", "").replace("_", " ").strip()

    if any(kw in lower for kw in ("fuel price", "diesel", "petrol", "fuel outlook")):
        # Fuel price report - extract country-level data
        articles.extend(
            _parse_fuel_price_pdf(full_text, clean_name, msg_date)
        )
    elif any(kw in lower for kw in ("acceptance", "closing date", "vessel", "eta")):
        # Vessel schedule - single consolidated article
        articles.append(Article(
            title=f"Vessel Schedule: {clean_name}",
            url="",
            source="Email Attachment",
            published_date=msg_date,
            raw_text=full_text[:2000],
            summary=f"Vessel acceptance and closing dates report. {_count_vessels(full_text)} vessels listed.",
            regions=_infer_regions_from_text(full_text),
            container_signal="general",
        ))
    else:
        # Generic PDF - treat as single intelligence article
        articles.append(Article(
            title=clean_name,
            url="",
            source="Email Attachment",
            published_date=msg_date,
            raw_text=full_text[:2000],
            summary=full_text[:300].replace("\n", " ").strip(),
            regions=_infer_regions_from_text(full_text),
            container_signal=_infer_signal(full_text),
        ))

    return articles


def _parse_fuel_price_pdf(
    text: str, title: str, msg_date: datetime
) -> List[Article]:
    """Parse a fuel price report PDF into per-country articles."""
    articles: List[Article] = []

    # Split by country names (common African countries in these reports)
    country_pattern = re.compile(
        r"^(South Africa|Namibia|Botswana|Zimbabwe|Mozambique|Malawi|"
        r"Zambia|DRC|Tanzania|Kenya|Uganda|Germany|Belgium|Italy|Spain|"
        r"Netherlands|France)\s*$",
        re.MULTILINE | re.IGNORECASE,
    )

    parts = country_pattern.split(text)
    # parts alternates: [preamble, country1, data1, country2, data2, ...]
    i = 1
    while i < len(parts) - 1:
        country = parts[i].strip()
        data = parts[i + 1].strip()[:500]
        if data:
            summary = f"{country}: {data[:250].replace(chr(10), ' ').strip()}"
            articles.append(Article(
                title=f"Fuel Price Update - {country}",
                url="",
                source="Fuel Price Report",
                published_date=msg_date,
                raw_text=data,
                summary=summary,
                regions=_country_to_regions(country),
                container_signal="general",
            ))
        i += 2

    # If no countries found, create a single article from the whole text
    if not articles:
        articles.append(Article(
            title=f"Fuel Price Report: {title}",
            url="",
            source="Email Attachment",
            published_date=msg_date,
            raw_text=text[:2000],
            summary=text[:300].replace("\n", " ").strip(),
            regions=[],
            container_signal="general",
        ))

    return articles


# ── XLSX attachment parsing ──────────────────────────────────────────────────

def _parse_xlsx_attachment(
    payload: bytes, filename: str, msg_date: datetime
) -> List[Article]:
    """Parse XLSX carrier advisory files into per-advisory articles."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed - skipping XLSX '%s'", filename)
        return []

    articles: List[Article] = []

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(payload)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
    except Exception as exc:
        logger.error("Failed to read XLSX '%s': %s", filename, exc)
        Path(tmp_path).unlink(missing_ok=True)
        return []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        carrier = sheet_name.strip().rstrip(".")

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        # Find header row (contains "Advisory Date" or "Advisory")
        header_idx = None
        for idx, row in enumerate(rows):
            row_str = " ".join(str(c or "") for c in row).lower()
            if "advisory date" in row_str or ("advisory" in row_str and "impact" in row_str):
                header_idx = idx
                break

        if header_idx is None:
            # Try WEC-style format (surcharge tables)
            articles.extend(_parse_surcharge_sheet(rows, carrier, msg_date))
            continue

        # Parse advisory rows
        for row in rows[header_idx + 1:]:
            cells = [str(c or "").strip() if c is not None else "" for c in row]
            if len(cells) < 3:
                continue

            # Try to find date, advisory title, impact, regions
            adv_date = _parse_cell_date(cells[0]) or msg_date
            advisory = cells[1] if len(cells) > 1 else ""
            impact = cells[2] if len(cells) > 2 else ""
            regions_text = cells[3] if len(cells) > 3 else ""

            if not advisory and not impact:
                continue
            # Skip blank/header rows
            if advisory.lower() in ("advisory", "") and not impact:
                continue

            title = f"{carrier}: {advisory[:80]}" if advisory else f"{carrier} Advisory"
            summary_parts = []
            if impact:
                summary_parts.append(impact[:300])
            if regions_text:
                summary_parts.append(f"Regions: {regions_text[:200]}")
            summary = " | ".join(summary_parts) or advisory

            articles.append(Article(
                title=title,
                url="",
                source=carrier,
                published_date=adv_date,
                raw_text=summary,
                summary=summary[:500],
                regions=_infer_regions_from_text(f"{advisory} {impact} {regions_text}"),
                container_signal=_infer_signal(f"{advisory} {impact}"),
            ))

    wb.close()
    Path(tmp_path).unlink(missing_ok=True)
    return articles


def _parse_surcharge_sheet(
    rows: list, carrier: str, msg_date: datetime
) -> List[Article]:
    """Parse WEC-style surcharge tables (POL/POD/surcharge amounts)."""
    articles: List[Article] = []

    for row in rows:
        cells = [str(c or "").strip() if c is not None else "" for c in row]
        row_text = " ".join(cells).lower()
        # Look for rows with port names and dollar amounts
        if "$" in row_text and any(kw in row_text for kw in ("mombasa", "dar", "djibouti", "antwerp", "rotterdam")):
            pol = cells[0] if cells else ""
            pod_region = cells[1] if len(cells) > 1 else ""
            summary = " | ".join(c for c in cells if c)[:400]
            articles.append(Article(
                title=f"{carrier}: Surcharge - {pol} → {pod_region}"[:100],
                url="",
                source=carrier,
                published_date=msg_date,
                raw_text=summary,
                summary=summary,
                regions=_infer_regions_from_text(f"{pol} {pod_region} {summary}"),
                container_signal="general",
            ))

    return articles


# ── Utility helpers ──────────────────────────────────────────────────────────

def _parse_email_date(date_str: str) -> datetime:
    """Parse email Date header into datetime."""
    if not date_str:
        return datetime.now(timezone.utc)
    try:
        from email.utils import parsedate_to_datetime
        return parsedate_to_datetime(date_str).astimezone(timezone.utc)
    except Exception:
        return datetime.now(timezone.utc)


def _parse_cell_date(value) -> Optional[datetime]:
    """Parse a date from an Excel cell value."""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
    try:
        from datetime import date as _date
        if isinstance(value, _date):
            return datetime(value.year, value.month, value.day, tzinfo=timezone.utc)
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(str(value).strip()[:19], fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _strip_email_noise(body: str) -> str:
    """Remove email signatures, disclaimers, and forwarding artifacts."""
    # Cut at common signature markers
    markers = [
        "Kind Regards", "Best Regards", "Best regards",
        "Disclaimer", "EXTERNAL SENDER",
        "This email is intended solely",
        "If our account details change",
        "________________________________",
    ]
    for marker in markers:
        idx = body.find(marker)
        if idx > 50:  # Keep at least 50 chars
            body = body[:idx]
            break

    # Remove forwarding headers
    body = re.sub(r"^From:.*?Subject:.*?\n", "", body, flags=re.DOTALL | re.MULTILINE)
    return body.strip()


def _extract_sender_name(sender: str) -> str:
    """Extract clean sender name from email From header."""
    if "<" in sender:
        name = sender.split("<")[0].strip().strip('"')
        if name:
            return name
    return sender.split("@")[0] if "@" in sender else sender


def _infer_signal(text: str) -> Optional[str]:
    """Infer container_signal from text content."""
    lower = text.lower()
    if any(kw in lower for kw in ("shortage", "depleted", "nil", "low", "critical", "suspend")):
        return "shortage"
    if any(kw in lower for kw in ("surplus", "relief", "normalized", "steady")):
        return "surplus"
    if any(kw in lower for kw in ("surcharge", "advisory", "update", "monitor", "fuel", "price")):
        return "general"
    return None


def _infer_regions_from_subject(subject: str) -> List[str]:
    """Infer region keys from email subject line."""
    lower = subject.lower()
    regions = []
    region_map = {
        "east_africa": ["kampala", "mombasa", "dar es salaam", "nairobi", "kenya", "uganda", "tanzania", "ethiopia", "djibouti"],
        "central_america": ["panama", "honduras", "nicaragua", "corinto", "colon"],
        "brazil": ["brazil", "santos", "paranagua"],
        "north_europe": ["rotterdam", "hamburg", "antwerp", "europe"],
        "vietnam": ["vietnam", "ho chi minh", "hai phong"],
        "middle_east": ["middle east", "gulf", "hormuz", "red sea", "dubai", "fujairah"],
    }
    for region_key, keywords in region_map.items():
        if any(kw in lower for kw in keywords):
            regions.append(region_key)
    return regions


def _infer_regions_from_text(text: str) -> List[str]:
    """Infer region keys from arbitrary text."""
    return _infer_regions_from_subject(text)


def _country_to_regions(country: str) -> List[str]:
    """Map a country name to region keys."""
    mapping = {
        "south africa": ["east_africa"],
        "namibia": ["east_africa"],
        "botswana": ["east_africa"],
        "zimbabwe": ["east_africa"],
        "mozambique": ["east_africa"],
        "malawi": ["east_africa"],
        "zambia": ["east_africa"],
        "drc": ["east_africa"],
        "tanzania": ["east_africa"],
        "kenya": ["east_africa"],
        "uganda": ["east_africa"],
        "germany": ["north_europe"],
        "belgium": ["north_europe"],
        "italy": ["north_europe"],
        "spain": ["north_europe"],
        "netherlands": ["north_europe"],
        "france": ["north_europe"],
    }
    return mapping.get(country.lower(), [])


def _count_vessels(text: str) -> int:
    """Count vessel entries in a schedule PDF."""
    # Each vessel has a 3-letter carrier code like MSC, CMA, MAE, PIL, etc.
    vessel_codes = re.findall(r"\b(MSC|CMA|MAE|PIL|COS|HLC|ONE|MES|ISS|SMK|AOL|SPM)\b", text)
    return len(set(re.findall(r"[A-Z]{4}-\d{4}", text)))
